# Instructions for running:
#
# 1. Create a virtual environment
#    `python3 -m venv venv`
# 2. Activate the virtual environment
#    `source venv/bin/activate`
# 3. Install dependencies
#    `pip install -U deep-translator openpyxl`
# 4. Change the input and output file names as required and run the script
#    `python3 translate.py`

from deep_translator import GoogleTranslator
from openpyxl import load_workbook


input_file = "/mnt/c/Users/augh/Desktop/translations.xlsx"
output_file = "/mnt/c/Users/augh/Desktop/translations_translated.xlsx"


class EnglishTranslator:
    @staticmethod
    def translate(text):
        return text


translators = {
    "English": EnglishTranslator,
    "Español": GoogleTranslator(source="en", target="es"),
    "Français": GoogleTranslator(source="en", target="fr"),
    "Dutch": GoogleTranslator(source="en", target="nl"),
    "German": GoogleTranslator(source="en", target="de"),
}

workbook = load_workbook(input_file)
worksheet = workbook.active

source = "Default"
source_index = 0
target = {}
for index, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1, values_only=True)):
    value = col[0].split()[0]
    target[index] = translators.get(value, value)

    if value == source:
        source_index = index

for row in worksheet.iter_rows(min_row=2):
    default = row[source_index].value

    for index, cell in enumerate(row):
        if cell.value is None:
            cell.value = target[index].translate(default)

    print(".", end="", flush=True)

workbook.save(output_file)
