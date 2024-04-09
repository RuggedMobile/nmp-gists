#v R.0.1

import argparse
from openpyxl import Workbook
from openpyxl import load_workbook

FIELD_TYPE_COLUMN = "B"
BACKGROUND_COLOR_COLUMN="N"
TITLE_STYLE_COLUMN="O"
HINT_STYLE_COLUMN="P"
OPTION_BACKGROUND_COLOR_COLUMN="Y"
BUTTON_STYLE_COLUMN="BB"
BUTTON_COLOR_COLUMN="BC"

DATA_NAME_COLUMN="C"

parser = argparse.ArgumentParser()
parser.add_argument("form")
args = parser.parse_args()

form1 = args.form
page_bg_color = "#F79A26"
page_ts = "#000000|large|bold"

section_bg_color = "#000000"
section_ts = "#F79A26|large|bold"
section_hs = "#F79A26|medium|bold"

general_bg_color = "#FFFFFF"
general_ts = "#000000|medium|bold"
general_hs = "#F79A26|medium|bold"

options_bg_color = "#606060"

buttons_style = "#000000*medium*bold|#FFFFFF*medium*bold|#FFFFFF*medium*bold|#FFFFFF*medium*bold|#FFFFFF*medium*bold"
buttons_color = "#FFFFFF|#000000|#000000|#000000|#000000"

wb = load_workbook(form1)
ws = wb.active

row = 2
while True:
    row = row + 1
    if (ws["{}{}".format(FIELD_TYPE_COLUMN,row)].value is None):
        break
    elif (ws["{}{}".format(FIELD_TYPE_COLUMN,row)].value == "Page"):
        ws["{}{}".format(BACKGROUND_COLOR_COLUMN, row)] = page_bg_color
        ws["{}{}".format(TITLE_STYLE_COLUMN, row)] = page_ts
    elif (ws["{}{}".format(FIELD_TYPE_COLUMN,row)].value == "Section"):
        if ("logo" in ws["{}{}".format(DATA_NAME_COLUMN,row)].value or "space" in ws["{}{}".format(DATA_NAME_COLUMN,row)].value): ## fixed DATA_NAME => DATA_NAME_COLUMN
            ws["{}{}".format(BACKGROUND_COLOR_COLUMN, row)] = "transparent"
        else:
            ws["{}{}".format(BACKGROUND_COLOR_COLUMN, row)] = section_bg_color
        ws["{}{}".format(TITLE_STYLE_COLUMN, row)] = section_ts
        ws["{}{}".format(HINT_STYLE_COLUMN, row)] = section_hs
    else:
        ws["{}{}".format(BACKGROUND_COLOR_COLUMN, row)] = general_bg_color
        ws["{}{}".format(TITLE_STYLE_COLUMN, row)] = general_ts
        ws["{}{}".format(HINT_STYLE_COLUMN, row)] = general_hs
        if (ws["{}{}".format(FIELD_TYPE_COLUMN,row)].value == "Choices"):
            ws["{}{}".format(OPTION_BACKGROUND_COLOR_COLUMN, row)] = options_bg_color
        elif (ws["{}{}".format(FIELD_TYPE_COLUMN,row)].value == "Action"):
            ws["{}{}".format(BUTTON_STYLE_COLUMN, row)] = buttons_style
            ws["{}{}".format(BUTTON_COLOR_COLUMN, row)] = buttons_color

form1 = form1.replace(".xlsx", "NEW.xlsx") ## needs to be different than wb or else error
print(form1)
wb.save(form1)
##problem when uploading (creates new row), need to delete in excel before import
